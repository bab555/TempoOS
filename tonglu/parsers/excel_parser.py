# Copyright (c) 2026 TempoOS Contributors. All Rights Reserved.

"""
Excel Parser — Extract text and structured data from .xlsx/.xls files.

Uses openpyxl (synchronous) wrapped in asyncio.to_thread.
"""

from __future__ import annotations

import asyncio
import logging
from typing import Any, Optional

import openpyxl

from tonglu.parsers.base import BaseParser, ParseResult

logger = logging.getLogger("tonglu.parsers.excel")


class ExcelParser(BaseParser):
    """Excel 文件解析器。"""

    async def parse(self, content_ref: str, **kwargs: Any) -> ParseResult:
        """Parse an Excel file, extracting all sheets as text + structured rows."""
        return await asyncio.to_thread(self._sync_parse, content_ref)

    def supports(self, file_name: Optional[str], source_type: str) -> bool:
        return bool(file_name) and file_name.lower().endswith((".xlsx", ".xls"))

    @staticmethod
    def _sync_parse(path: str) -> ParseResult:
        """Synchronous Excel parsing — runs in thread pool."""
        wb = openpyxl.load_workbook(path, data_only=True)
        text_parts: list[str] = []
        tables: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(row_data)
                text_parts.append(" | ".join(row_data))
            tables.append({"sheet": sheet_name, "rows": rows})

        total_rows = sum(len(t["rows"]) for t in tables)
        logger.debug(
            "Excel parsed: %s — %d sheets, %d rows",
            path, len(wb.sheetnames), total_rows,
        )

        return ParseResult(
            text="\n".join(text_parts),
            metadata={
                "sheets": wb.sheetnames,
                "total_rows": total_rows,
            },
            tables=tables,
        )
